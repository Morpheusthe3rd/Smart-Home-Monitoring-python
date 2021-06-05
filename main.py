import os
import random
import time
import numpy
import queue
from openpyxl import Workbook
import serial
import json

dataDict = {  # This places the readings into a dictionary
    "temperature": 0,
    "HVACStatus": 0,
    "currentData": [0,0,0,0,0,0,0,0,0,0,0,0,0]
}

def current_test():

    #The current test block initializes a serial connection with the Arduino, sends the initializing j across and tokenizes the recieved data.

    ser = serial.Serial("/dev/ttyACM0", 9600) #The channel for the Arduino will be ttyACM0
    print("serial established")

    msg = "j\n"

    ser.write(msg.encode())   
    
    serRead = ser.readline().decode()
    #with open(serRead) as sr:
    #dict = json.loan(sr)
    print("dict recieved")
    print(dict);
    currentArray = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

    # This block tokenizes the string, extracting the (float) readings
    strArray = tokenize(dict, ", ")
    i=0
    for(String e : strArray):
        currentArray[i] = (float)tokenize(e, ": ")
        i++

    sum = 0
    for e in range (0, 11): # Performs a sum to provide total energy usage
        sum += currentArray[e]
    currentArray[12] = sum
    
    return currentArray

def tenMinutePause(prevTime):
    time.sleep(600-prevTime) # This pauses the program for 10 minutes

def initializeDataArray():  # This helper function initializes a model for the data storage.
    weeksByHour = queue.Queue(168)
    monthByDay = queue.Queue(35)
    yearsByMonth = queue.Queue(60)
    mainDict = {
        "dayByMinute": [],
        "weekByHour": weeksByHour,
        "monthByDays": monthByDay,
        "yearsByMonth": yearsByMonth
    }
    # This loop will expand the lower array to the size needed
    for i in range (145):
        mainDict["dayByMinute"].append(dataDict)

    return mainDict

def updateDataArray(currentMinuteIndex, dataArray, updateMinuteBool, updateHourBool, updateDayBool, updateMonthBool):

    #some testing prints to ensure accuracy

    #print("currentMinuteIndex: {0}".format(currentMinuteIndex))
    #print("updateMinuteBool: {0}".format(updateMinuteBool))
    #print("updateHourBool: {0}".format(updateHourBool))
    #print("updateDayBool: {0}".format(updateDayBool))
    #print("updateMonthBool: {0}\n\n".format(updateMonthBool))

    # This string of if statements checks which sections need updated, and create the necessary memory

    if(updateMonthBool):
        ave = 0
        for i in range(35):
            ave += dataArray["monthByDay"][i]

        ave /= 35

        if (dataArray["yearsByMonth"].full()):
            dataArray["yearsByMonth"].get()
            dataArray["yearsByMonth"].put(ave)
        else:
            dataArray["yearsByMonth"].put(ave)

    if(updateDayBool):
        ave = 0
        for i in range(24):
            ave += dataArray["weekByHour"][i]["currentData"][12]

        ave /= 24

        if (dataArray["monthByDay"].full()):
            dataArray["monthByDay"].get()
            dataArray["monthByDay"].put(ave)
        else:
            dataArray["monthByDay"].put(ave)

    if(updateHourBool):
        averages = []
        ave = 0

        for j in range(12):
            for i in range (6):
                #print(dataArray["dayByMinute"][j]["currentData"][1])
                #print("i, j = {0}, {1}".format(i, j))
                #print("value at i, j = {0}".format(dataArray["dayByMinute"][currentMinuteIndex-i]["currentData"][j]))
                ave = ave + dataArray["dayByMinute"][currentMinuteIndex-i]["currentData"][j]
            #print(ave)
            ave /= 6
            averages.append(ave)

        dataDict["currentData"] = averages

        if(dataArray["weekByHour"].full()):
            dataArray["weekByHour"].get()
            dataArray["weekByHour"].put(dataDict)
        else:
            dataArray["weekByHour"].put(dataDict)

    if(updateMinuteBool):
        dataArray["dayByMinute"] = []
        for i in range(145):
            dataArray["dayByMinute"].append(dataDict)

def excelUpload(dataStruct): # This, what is essentially a printing function, will linearly process the dictionary and save all values to the sheet
    wb = Workbook()
    ws = wb.active

    i = 0
    # 10 minute readings, circuit(column) 1
    for e in dataStruct["dayByMinute"]:
        #print("e: {0}".format(e))
        #print("e[currentData]".format(e["currentData"]))
        #print("e[currentData][0] = {0}".format(e["currentData"][0]))

        cellID = 'A' + str(2+i)

        ws[cellID] = e["currentData"][0]
        i += 1
    i = 0
    # 10 minute readings, circuit(column) 2
    for e in dataStruct["dayByMinute"]:
        cellID = 'B' + str(2 + i)

        ws[cellID] = e["currentData"][1]
        i += 1
    i = 0
    # 10 minute readings, circuit(column) 3
    for e in dataStruct["dayByMinute"]:
        cellID = 'C' + str(2 + i)

        ws[cellID] = e["currentData"][2]
        i += 1
    i = 0
    # 10 minute readings, circuit(column) 4
    for e in dataStruct["dayByMinute"]:
        cellID = 'D' + str(2 + i)

        ws[cellID] = e["currentData"][3]
        i += 1
    i = 0
    # 10 minute readings, circuit(column) 5
    for e in dataStruct["dayByMinute"]:
        cellID = 'E' + str(2 + i)

        ws[cellID] = e["currentData"][4]
        i += 1
    i = 0
    # 10 minute readings, circuit(column) 6
    for e in dataStruct["dayByMinute"]:
        cellID = 'F' + str(2 + i)

        ws[cellID] = e["currentData"][5]
        i += 1
    i = 0
    # 10 minute readings, circuit(column) 7
    for e in dataStruct["dayByMinute"]:
        cellID = 'G' + str(2 + i)

        ws[cellID] = e["currentData"][6]
        i += 1
    i = 0
    # 10 minute readings, circuit(column) 8
    for e in dataStruct["dayByMinute"]:
        cellID = 'H' + str(2 + i)

        ws[cellID] = e["currentData"][7]
        i += 1
    i = 0
    # 10 minute readings, circuit(column) 9
    for e in dataStruct["dayByMinute"]:
        cellID = 'I' + str(2 + i)

        ws[cellID] = e["currentData"][8]
        i += 1
    i = 0
    # 10 minute readings, circuit(column) 10
    for e in dataStruct["dayByMinute"]:
        cellID = 'J' + str(2 + i)

        ws[cellID] = e["currentData"][9]
        i += 1
    i = 0
    # 10 minute readings, circuit(column) 11
    for e in dataStruct["dayByMinute"]:
        cellID = 'K' + str(2 + i)

        ws[cellID] = e["currentData"][10]
        i += 1
    i = 0
    # 10 minute readings, circuit(column) 12
    for e in dataStruct["dayByMinute"]:
        cellID = 'L' + str(2 + i)

        ws[cellID] = e["currentData"][11]
        i += 1
    i = 0
    # 10 minute readings, total(column) 13
    for e in dataStruct["dayByMinute"]:
        cellID = 'M' + str(2 + i)
        #print("e['currentData']: {0}".format(e["currentData"]))
        ave = 0
        for x in range(12):
            ave += e["currentData"][x]

        ws[cellID] = ave
        i += 1
    i = 0

#=============================================

    for e in dataStruct["monthByDays"].queue:
        cellID = 'ac' + str(2 + i)
        print("test")
        #print("e: {0}".format(e))
        ws[cellID] = e["currentData"][0]
        i += 1
    i = 0
    for e in dataStruct["monthByDays"].queue:
        cellID = 'ad' + str(2 + i)
        ws[cellID] = e["currentData"][1]
        i += 1
    i = 0
    for e in dataStruct["monthByDays"].queue:
        cellID = 'ae' + str(2 + i)
        ws[cellID] = e["currentData"][2]
        i += 1
    i = 0
    for e in dataStruct["monthByDays"].queue:
        cellID = 'af' + str(2 + i)
        ws[cellID] = e["currentData"][3]
        i += 1
    i = 0
    for e in dataStruct["monthByDays"].queue:
        cellID = 'ag' + str(2 + i)
        ws[cellID] = e["currentData"][4]
        i += 1
    i = 0
    for e in dataStruct["monthByDays"].queue:
        cellID = 'ah' + str(2 + i)
        ws[cellID] = e["currentData"][5]
        i += 1
    i = 0
    for e in dataStruct["monthByDays"].queue:
        cellID = 'ai' + str(2 + i)
        ws[cellID] = e["currentData"][6]
        i += 1
    i = 0
    for e in dataStruct["monthByDays"].queue:
        cellID = 'aj' + str(2 + i)
        ws[cellID] = e["currentData"][7]
        i += 1
    i = 0
    for e in dataStruct["monthByDays"].queue:
        cellID = 'ak' + str(2 + i)
        ws[cellID] = e["currentData"][8]
        i += 1
    i = 0
    for e in dataStruct["monthByDays"].queue:
        cellID = 'al' + str(2 + i)
        ws[cellID] = e["currentData"][9]
        i += 1
    i = 0
    for e in dataStruct["monthByDays"].queue:
        cellID = 'am' + str(2 + i)
        ws[cellID] = e["currentData"][10]
        i += 1
    i = 0
    for e in dataStruct["monthByDays"].queue:
        cellID = 'an' + str(2 + i)
        ws[cellID] = e["currentData"][11]
        i += 1
    i = 0
    for e in dataStruct["monthByDays"].queue:
        ave = 0
        cellID = 'ao' + str(2 + i)
        for x in range(12):
            ave += e["currentData"][x]

        ws[cellID] = ave
        i += 1
    i = 0

    # =============================================

    for e in dataStruct["weekByHour"].queue:
        cellID = 'O' + str(2 + i)
        #print("e: {0}".format(e))
        ws[cellID] = e["currentData"][0]
        i += 1
    i = 0
    for e in dataStruct["weekByHour"].queue:
        cellID = 'P' + str(2 + i)
        ws[cellID] = e["currentData"][1]
        i += 1
    i = 0
    for e in dataStruct["weekByHour"].queue:
        cellID = 'q' + str(2 + i)
        ws[cellID] = e["currentData"][2]
        i += 1
    i = 0
    for e in dataStruct["weekByHour"].queue:
        cellID = 'r' + str(2 + i)
        ws[cellID] = e["currentData"][3]
        i += 1
    i = 0
    for e in dataStruct["weekByHour"].queue:
        cellID = 's' + str(2 + i)
        ws[cellID] = e["currentData"][4]
        i += 1
    i = 0
    for e in dataStruct["weekByHour"].queue:
        cellID = 't' + str(2 + i)
        ws[cellID] = e["currentData"][5]
        i += 1
    i = 0
    for e in dataStruct["weekByHour"].queue:
        cellID = 'u' + str(2 + i)
        ws[cellID] = e["currentData"][6]
        i += 1
    i = 0
    for e in dataStruct["weekByHour"].queue:
        cellID = 'v' + str(2 + i)
        ws[cellID] = e["currentData"][7]
        i += 1
    i = 0
    for e in dataStruct["weekByHour"].queue:
        cellID = 'w' + str(2 + i)
        ws[cellID] = e["currentData"][8]
        i += 1
    i = 0
    for e in dataStruct["weekByHour"].queue:
        cellID = 'x' + str(2 + i)
        ws[cellID] = e["currentData"][9]
        i += 1
    i = 0
    for e in dataStruct["weekByHour"].queue:
        cellID = 'y' + str(2 + i)
        ws[cellID] = e["currentData"][10]
        i += 1
    i = 0
    for e in dataStruct["weekByHour"].queue:
        cellID = 'z' + str(2 + i)
        ws[cellID] = e["currentData"][11]
        i += 1
    i = 0
    for e in dataStruct["weekByHour"].queue:
        ave = 0
        cellID = 'aa' + str(2 + i)
        for x in range(12):
            ave += e["currentData"][x]

        ws[cellID] = ave
        i += 1
    i = 0

#====================================================================================

    for e in dataStruct["yearsByMonth"].queue:
        cellID = 'aq' + str(2+i)
        ws[cellID] = e
        i += 1
    i = 0
    wb.save('SenDesTest.xlsx')

if __name__ == '__main__':  # The main function runs in an endless while loop, taking readings, updating memory and uploading to the excel spreadsheet every 10 minutes.

    minutesPerDay = 144     #This is represented as /10
    minutesPerHour = 60
    hoursPerWeek = 24 * 7
    daysPerWeek = 7
    daysPerMonth = 35
    weeksPerMonth = 5

    minutesPassed = 0
    hourCount = 0
    dayCount = 0
    monthCount = 0
    yearCount = 0

    looptime = 0

    fullDataArray = initializeDataArray()

    while(True):

        tenMinutePause(looptime) # this pauses the loop

        startTime = time.perf_counter() # sets up a timer for this whole loop



        TempData, HVACStatus = T9_Interface_test()      # These lines draw the data from the subroutines.
        currentData = current_test()

        if HVACStatus:
            HVACStatus = 10
        else:
            HVACStatus = 0

        dataDict = {  # This places the readings into a dictionary
            "temperature": TempData,
            "HVACStatus": HVACStatus,
            "currentData": currentData
        }
        #print(len(fullDataArray["dayByMinute"]))

        print("minutesPassed: {0}".format(minutesPassed))
        print("hourCount: {0}".format(hourCount))
        print("monthCount: {0}".format(monthCount))
        print("yearCount: {0}".format(yearCount))

        fullDataArray["dayByMinute"][minutesPassed] = dataDict

        #print(fullDataArray)
        #print((minutesPassed >= minutesPerDay))

        updateDataArray(minutesPassed, fullDataArray, (minutesPassed >= minutesPerDay), (hourCount==60), (dayCount == 24), (monthCount == 35))
        excelUpload(fullDataArray)

        if (minutesPassed==144):
            minutesPassed = 0
        if (hourCount==60):
            hourCount = 0
            dayCount += 1

        if (dayCount == 24):
            dayCount = 0
            monthCount += 1

        if (monthCount == 35):
            monthCount = 0
            yearCount += 1
        if yearCount == 12:
            break

        #print(TempData)
        #print(HVACStatus)
        #print(currentData)

        minutesPassed += 1
        hourCount += 10
        looptime = time.perf_counter() - startTime
        #print("This loop took: {0}".format(looptime))
